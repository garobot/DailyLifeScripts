# import the necessary files
import cv2 as cv
import openpyxl
from PIL import Image

template_path = r"path to your certificate template"

# Excel file containing names of participants
participants_list = r"path to your participants excel file"

# Path to store the certificates
output_path = r"path to save the certificate"

# Function to convert image to pdf
def convert_to_pdf(path):
    certi = Image.open(path)
    C = certi.convert("RGB")
    pdf_path = output_path + "/" + certi_name + "_certificate.pdf"
    C.save(pdf_path)


# Setting the font, font size and font colour
font = cv.FONT_HERSHEY_SIMPLEX
font_size = 2
font_color = (0, 0, 0)

# Coordinates on the certificate where name is to be printed
x_pos, y_pos = 15, 7

# loading the participants.xlsx workbook and grabbing the active sheet
obj = openpyxl.load_workbook(participants_list)
sheet = obj.active

# Finding the number of rows in sheet
rows = sheet.max_row

# Printing certificate for each participant
for i in range(1, rows + 1):

    # i'th participant
    get_name = sheet.cell(row=i, column=1)
    certi_name = get_name.value

    # Read the certificate template
    img = cv.imread(template_path)

    # The size of name to be printed
    text_size = cv.getTextSize(certi_name, font, font_size, 10)[0]

    # The (x, y) coordinates of the name to be printed
    text_x = int((img.shape[1] - text_size[0]) / 2 + y_pos)
    text_y = int((img.shape[0] + text_size[1]) / 2 - x_pos)
    cv.putText(img, certi_name, (text_x, text_y), font, font_size, font_color, 10)

    # Save the certificate at the specific path
    certi_path = output_path + "/" + certi_name + "_certificate" + ".png"
    cv.imwrite(certi_path, img)
    convert_to_pdf(certi_path)
